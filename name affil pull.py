import re
import spacy
import nltk.tokenize
from nltk.tokenize import sent_tokenize
#nltk.download('punkt')
import json
import csv
import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook
#initialize spacy
nlp = spacy.load("en_core_web_sm")

sentlist = []
sententlist= []
namelist= []

jsonfile= open("jsonfilename", encoding = 'UTF-8')
jsondata= json.load(jsonfile)

for i in jsondata:
    if i.get('sentence_match') is not None:
        sentlist.append(i.get('sentence_match'))
    elif i.get('sent_match') is not None:
        sentlist.append(i.get('sent_match'))
    else:
        print("failure to get sentence")
        continue

for i, sent in enumerate(sentlist):
    if sent is not None:
        nlpsent = nlp(sent)
    else:
        print('rip this one failed' + str(i))
        continue
    for ent in nlpsent.ents:
        if ent.label_ =='PERSON':
            result= (ent.text)
            #write person entities to namelist
            namelist.append(result)
            #write anything in namelist to sentlist, which will be the lists of lists
            #sententlist.append(namelist)
        else:
            result =('no person')
            namelist.append(result)
            continue
    if re.search('at the (.*),', sent):
        school = re.search('at the (.*),', sent)
        school = school.group()
        school = re.split('at |the |and', school)
        reformlist = list(filter(None, school))
        namelist.extend(reformlist)
        sententlist.append(namelist)
        namelist = []
    elif re.search('at the (.*) and', sent):
        school = re.search('at the (.*) and', sent)
        school = school.group()
        school = re.split('at |the |and', school)
        reformlist = list(filter(None, school))
        namelist.extend(reformlist)
        sententlist.append(namelist)
        namelist = []
    elif re.search('at the(.*).', sent):
        school= re.search('at the(.*).', sent)
        school = school.group()
        school = re.split('at |the |and', school)
        reformlist = list(filter(None, school))
        namelist.extend(reformlist)
        sententlist.append(namelist)
        namelist = []
    elif re.search('at(.*),', sent):
        school = re.search ('at(.*),', sent)
        school = school.group()
        school = re.split('at |the |and', school)
        reformlist = list(filter(None, school))
        namelist.extend(reformlist)
        sententlist.append(namelist)
        namelist = []
    elif re.search('at(.*) and', sent):
        school = re.search ('at(.*) and', sent)
        school = school.group()
        school = re.split('at |the |and', school)
        reformlist = list(filter(None, school))
        namelist.extend(reformlist)
        sententlist.append(namelist)
        namelist = []
    elif re.search('at the(.*)and', sent):
        school = re.search('at the(.*)and', sent)
        school = school.group()
        school = re.split('at |the |and', school)
        reformlist = list(filter(None, school))
        namelist.extend(reformlist)
        sententlist.append(namelist)
        namelist = []
    else:
        reformlist = ['none sorry']
        namelist.extend(reformlist)
        sententlist.append(namelist)
        namelist = []
        continue
        print('yeehaw')

#where the sentence is col 1, and acts as an ID
#and each object in the list is in the same row, in the next column
#so: (sent1, [joe bob, uwm]) sent1 is cell: col1xrow1; joe bob is cell col2xrow1; and uwm is col3xrow1
wb = load_workbook("file path name")
sheet1=wb.active

for i, s in enumerate(sentlist):
    sheet1.cell(i+1, 1).value = s
for i, q in enumerate(sententlist):
    for j, s in enumerate(q):
        sheet1.cell(i+1, j+2).value = s

wb.save("file path name")





